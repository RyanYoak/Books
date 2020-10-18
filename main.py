import books
import Book
import openpyxl
import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException

book_data = []
sleep_time = 5
months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
test_books = ["The Locus Awards: Thirty Years of the Best in Science Fiction and Fantasy "]
wb = openpyxl.Workbook()


browser = webdriver.Firefox()
browser.get('https://www.goodreads.com/search?q=first&qid=')

def find_title():
    try:
        title_element = browser.find_element_by_id("bookTitle")
    except NoSuchElementException:
        return "Unknown"
    title_text = title_element.text
    return title_text

def find_author():
    try:
        temp = browser.find_element_by_xpath("//span[@itemprop='author']")
        more_authors_item = temp.find_element_by_xpath(".//span[@class='toggleLink']")
        more_authors_item.click()
    except NoSuchElementException:
        x = 1
    try:
        author_elements = browser.find_elements_by_xpath("//span[@itemprop='name']")
    except NoSuchElementException:
        return "Unknown"
    authors = ""
    authors += author_elements[0].text
    for i in range(1, len(author_elements)):
        authors = authors + ", " + author_elements[i].text
    return authors

def find_pages():
    try:
        pages_element = browser.find_element_by_xpath("//span[@itemprop='numberOfPages']").text
    except NoSuchElementException:
        return "Unknown"
    pages_element = pages_element[:-6]
    return int(pages_element)

def find_genres():
    genres_string = ""
    try:
        genres = browser.find_elements_by_xpath("//div[@class='elementList ']")
    except NoSuchElementException:
        return "Unknown"
    max_genres = 5
    if len(genres) < 5:
        max_genres = len(genres)
    for i in range(max_genres):
        sub_genres = genres[i].find_elements_by_xpath(".//a[@class='actionLinkLite bookPageGenreLink']")
        genres_string += sub_genres[0].text
        if len(sub_genres) > 1:
            for i in range(1, len(sub_genres)):
                genres_string = genres_string + " > " + sub_genres[i].text
        genres_string += ", "
    return genres_string[:-2]

def find_published():
    date_string = ""
    month = None
    day = None
    year = None
    try:
        details = browser.find_element_by_xpath("//div[@id='details']")
        publishing_details = details.find_elements_by_xpath(".//div[@class='row']")[1]
    except NoSuchElementException:
        return "Unknown"

    full_text = publishing_details.text
    text_list = full_text.split()

    if "first published" in full_text:
        year = text_list[len(text_list) - 1][:-1]
        try:
            day = int(text_list[len(text_list) - 2][:-2])
        except ValueError:
            day = 1
        if text_list[len(text_list) - 2] in months or text_list[len(text_list) - 3] in months:
            if text_list[len(text_list) - 2] in months:
                month = text_list[len(text_list) - 2]
            else:
                month = text_list[len(text_list) - 3]
        else:
            month = "January"
    else:
        if text_list[1] in months:
            month = text_list[1]
        else:
            month = "January"
            day = 1
            year = text_list[1]
            return str(months.index(month) + 1) + "/" + str(day) + "/" + year
        if "st" in text_list[2] or "nd" in text_list[2] or "th" in text_list[2]:
            day = text_list[2][:-2]
            year = text_list[3]
        else:
            day = 1
            year = text_list[2]

    return str(months.index(month) + 1) + "/" + str(day) + "/" + year

def find_rating():
    try:
        rating_element = browser.find_element_by_xpath("//span[@itemprop='ratingValue']")
    except NoSuchElementException:
        return "Unknown"
    return rating_element.text

def find_series():
    try:
        series_element = browser.find_element_by_xpath("//h2[@id='bookSeries']")
    except NoSuchElementException:
        return "Unknown"
    series_text = series_element.text
    contains_series_number = 0
    if "#" in series_text:
        contains_series_number = 1
    if len(series_text) > 0:
        series_text_list = series_text.split()
        series_text_list[0] = series_text_list[0][1:]
        series_text = ""
        series_text += series_text_list[0]
        for i in range(1, len(series_text_list) - contains_series_number):
            series_text = series_text + " " + series_text_list[i]
        if not bool(contains_series_number):
            series_text = series_text[:-1]
        return series_text
    else:
        return ""

def get_to_page(book):
    search_bar = browser.find_element_by_name("q")
    search_bar.send_keys(book)
    search_bar.send_keys(Keys.RETURN)
    time.sleep(sleep_time)
    first_item = browser.find_element_by_class_name("bookTitle")
    browser.execute_script("arguments[0].click();",first_item)
    time.sleep(sleep_time)

def setup_sheet():
    ws = wb.active
    ws.title = "All Books"
    ws['A1'] = "Title"
    ws['B1'] = "Author"
    ws['C1'] = "Pages"
    ws['D1'] = "Published"
    ws['E1'] = "Rating"
    ws['F1'] = "Series"
    ws['G1'] = "Genre"

for book in books.book_titles:
#for book in test_books:
    get_to_page(book)
    temp = None
    print("Now getting: " + book)
    try:
        temp = Book.Book(find_title(), find_author(), find_pages(), find_series(), find_genres(), find_published(), find_rating())
    except:
        sleep_time = 10
        get_to_page(book)
        temp = Book.Book(find_title(), find_author(), find_pages(), find_series(), find_genres(), find_published(), find_rating())
        sleep_time = 5

    print(temp)
    book_data.append(temp)
browser.close()

row = 2
setup_sheet()
ws = wb.active
for book in book_data:
    ws["A{0}".format(row)] = book.title
    ws["B{0}".format(row)] = book.author
    ws["C{0}".format(row)] = book.pages
    ws["D{0}".format(row)] = book.published
    ws["E{0}".format(row)] = book.rating
    ws["F{0}".format(row)] = book.series
    ws["G{0}".format(row)] = book.genre
    row = row + 1

wb.save('All Books.xlsx')
